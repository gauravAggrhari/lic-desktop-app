import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import re
import copy
import datetime
import shutil
from datetime import date
from dateutil.relativedelta import relativedelta
from Mail import mail

class excel(object):
    def __init__(self):
        # providing filename of excel with full path
        os.makedirs("C:/Users/AJAY LIC/Documents/LIC", exist_ok=True)
        self.filename = "C:/Users/AJAY LIC/Documents/LIC/Client Details.xlsx"
        # check for file, workbook and worksheet existence
        if os.path.isfile(self.filename):
            self.wb = load_workbook(self.filename)
            if 'Details' in self.wb.sheetnames:
                self.ws = self.wb.active
                self.row_count = int(self.ws.max_row)
            # creating work sheet if not present
            else:
                self.ws = self.wb.create_sheet(index=0, title='Details')
                self.headers()
                self.row_count = int(self.ws.max_row)
                self.wb.save(self.filename)
        # creating workbook and worksheet if not present
        else:
            self.wb = Workbook()
            self.ws = self.wb.create_sheet(index=0, title='Details')
            self.headers()
            self.row_count = int(self.ws.max_row)
            self.wb.save(self.filename)

    # appending column headings to the worksheet
    def headers(self):
        self.header = ['S.No', 'Name', 'Address', 'Mobile', 'Policy Number', 'DOC', 'FUP', 'Period', \
                       'Premium', 'Daily Collection', 'Target', 'Balance']
        self.ws.append(self.header)
        self.wb.save(self.filename)

    # adding new customer to the worksheet
    def add_client(self, entry):
        self.row_count = int(self.ws.max_row)
        # inserting serial number to the new entry
        entry.insert(0, self.row_count)
        """
        # calculating fup
        doc = entry[5].split('-')
        if entry[6] == "Yearly":
            doc[2] = int(doc[2])+1
        elif entry[6] == "Half-Yearly":
            doc[1] = int(doc[1]) + 6
        elif entry[6] == "Quarterly":
            doc[1] = int(doc[1]) + 3
        else:
            doc[1] = int(doc[1]) + 1

        doc = doc[0]+"-"+str(doc[1])+"-"+str(doc[2])
        # inserting fup to the new entry
        entry.insert(6, doc)
        """
        # inserting initial Daily Collection as 0 to the new entry
        entry.insert(9, 0)
        # inserting balance = premium to the new entry
        entry.append(entry[8])
        for row in range(2, self.ws.max_row+1):
            policy = self.ws.cell(row=row, column=5).value
            if str(policy) == str(entry[4]):
                return "Policy Number already exists. Please enter a unique number"
        self.ws.append(entry)
        self.wb.save(self.filename)
        # checking for increment in max_row to confirm update of worksheet
        if self.ws.max_row == self.row_count+1:
            return 1
        else:
            return 0

    # updating customer details to the worksheet
    def update_client(self, update_entry, sl_no):
        try:
            # index of the updated list
            index = 0

            # finding the row value of the updated entry
            for row in range(2, self.ws.max_row+1):
                cell_number = self.ws.cell(row=row, column=1)
                if str(sl_no) == str(cell_number.value).lower():
                    row_no = row
                    break

            # getting the value of premium, daily collection before update to calculate balance
            self.premium_before = int(self.ws.cell(row=int(row_no), column=9).value)
            #daily_collection_before = int(self.ws.cell(row=int(row_no), column=10).value)

            # updating the changed values to worksheet
            for i in range(2, 12):
                self.ws.cell(row=int(row_no), column=i).value = update_entry[index]
                index += 1

            # getting the values of updated premium, balance and daily collection to calculate balance
            self.updated_premium = int(self.ws.cell(row=int(row_no), column=9).value)
            #daily_collection = int(self.ws.cell(row=int(row_no), column=10).value)
            #updated_daily_collection = daily_collection_before + daily_collection
            # commenting the below line as it keeps incrementing the daily collection value
            #self.ws.cell(row=int(row_no), column=10).value = updated_daily_collection
            balance = self.ws.cell(row=int(row_no), column=12).value
            balance = int(balance)+(int(self.updated_premium)-int(self.premium_before))
            #balance = int(balance)-int(daily_collection)
            self.ws.cell(row=row_no, column=12).value = balance

            # calling update_daily_client function to create a new column with today's date as heading
            #self.update_daily_client()

            # updating the new column and corresponding entry row with today's daily collection
            #self.ws.cell(row=row_no, column=self.ws.max_column).value = updated_daily_collection

            # calling quarterly to check for 90 days and do a clean up of daily collection columns
            self.quarterly()
            self.wb.save(self.filename)
        except Exception as e:
            return 0

    # creating new column with the current date whenever update_client is called
    def update_daily_client(self):
        self.date = datetime.datetime.now().strftime("%d/%m/%Y")
        cell_date = self.ws.cell(row=1, column=self.ws.max_column)
        if str(cell_date.value) == str(self.date):
            pass
        else:
            self.ws.cell(row=1, column=int(self.ws.max_column)+1).value = self.date

        self.wb.save(self.filename)

    # deleting the customer entry selected on GUI, from worksheet
    def delete_client(self, serial):
        self.row_count = self.ws.max_row
        for row in range(2, self.ws.max_row+1):
            cell_number = self.ws.cell(row=row, column=1)
            if str(serial) == str(cell_number.value).lower():
                self.ws.delete_rows(row)
                break

        for del_row in range(1, self.ws.max_row):
            self.ws.cell(row=int(del_row)+1, column=1).value = int(del_row)

        self.wb.save(self.filename)
        if self.ws.max_row == self.row_count-1:
            return 1
        else:
            return 0

    def view_today(self):
        entry = list()
        details = list()
        total = 0
        if self.ws.cell(row=1, column=self.ws.max_column).value == "Balance":
            today_excel = datetime.datetime(year=1800, month=4, day=20).strftime("%d/%m/%Y")
        else:
            today_excel = self.ws.cell(row=1, column=self.ws.max_column).value
        try:
            today_split = today_excel.split('/')
            today = self.datetime(today_split)
            date_now = datetime.datetime.now().strftime("%d/%m/%Y")
            now_split = date_now.split("/")
            now = self.datetime(now_split)
            for i in range(2, self.ws.max_row + 1):
                cell_s_no = self.ws.cell(row=i, column=1).value
                entry.append(cell_s_no)
                cell_name = self.ws.cell(row=i, column=2).value
                entry.append(cell_name)
                cell_address = self.ws.cell(row=i, column=3).value
                entry.append(cell_address)
                cell_policy = self.ws.cell(row=i, column=5).value
                entry.append(cell_policy)
                if now == today:
                    cell_collection = self.ws.cell(row=i, column=self.ws.max_column).value
                    entry.append(cell_collection)
                    try:
                        total = total + int(self.ws.cell(row=i, column=self.ws.max_column).value)
                    except TypeError:
                        total = total + 0
                elif now > today:
                    cell_collection = 0
                    entry.append(cell_collection)
                details.append(copy.deepcopy(entry))
                del entry[:]
        except:
            pass
        return details, total

    # returns the entries searched for, to the GUI
    def view_client(self, name, policy):  # name - input from GUI to search for the entry in worksheet
        flag = 0
        data = list()
        index = 1
        index_match = list()
        match = list()
        name = name.lower()

        # finding row numbers in the worksheet, which matches with the search category
        for row in range(2, self.ws.max_row+1):
            index += 1
            if name == "" and policy == "":
                cell_name = self.ws.cell(row=row, column=2)
                if re.search(name, str(cell_name.value).lower()):
                    index_match.append(index)
                    flag = 1
            elif name == "" and policy != "":
                cell_policy = self.ws.cell(row=row, column=5)
                if re.search(policy, str(cell_policy.value).lower()):
                    index_match.append(index)
                    flag = 1
            elif name != "" and policy == "":
                cell_name = self.ws.cell(row=row, column=2)
                if re.search(name, str(cell_name.value).lower()):
                    index_match.append(index)
                    flag = 1
            else:
                cell_name = self.ws.cell(row=row, column=2)
                cell_policy = self.ws.cell(row=row, column=5)
                if re.search(name, str(cell_name.value).lower()):
                    if re.search(policy, str(cell_policy.value).lower()):
                        index_match.append(index)
                        flag = 1

        if flag == 1:
            for index_row in index_match:
                for i in range(1, self.ws.max_column + 1):
                    cell_obj = self.ws.cell(row=index_row, column=i)
                    match.append(cell_obj.value)

                data.append(copy.deepcopy(match))
                del match[:]

            return data
        
        else:
            return "No Customer details available with the given data"

    def backup(self):
        os.makedirs("C:/Users/AJAY LIC/Documents/Backup", exist_ok=True)
        file = "C:/Users/AJAY LIC/Documents/Backup/Backup_"+str(datetime.datetime.now().strftime("%d_%m_%Y")+".xlsx")
        shutil.copyfile(self.filename, file)

    def quarterly(self):
        '''
        head_13 = self.ws.cell(row=1, column=13)
        self.toaddr = "gauavaggrahari@@gmail.com"
        try:
            head_split = head_13.value.split("/")
            if int(head_split[0][:1]) == 0:
                if int(head_split[1][:1]) == 0:
                    self.quarter = datetime.datetime(year=int(head_split[2]), month=int(head_split[1][1:]), day=int(head_split[0][1:])) + relativedelta(months=+3, days=+1)
                    #self.quarter = datetime.datetime.now()
                    if self.quarter.strftime("%d/%m/%Y") <= self.date:
                        flag = mail(self.toaddr)
                        if flag == 1:
                            self.ws.delete_cols(13, self.ws.max_column-13)
                        else:
                            self.backup()
                            self.ws.delete_cols(13, self.ws.max_column - 13)
                else:
                    self.quarter = datetime.datetime(year=int(head_split[2]), month=int(head_split[1]), day=int(head_split[0][1:])) + relativedelta(months=+3, days=+1)
                    if self.quarter.strftime("%d/%m/%Y") <= self.date:
                        flag = mail(self.toaddr)
                        if flag == 1:
                            self.ws.delete_cols(13, self.ws.max_column - 13)
                        else:
                            self.backup()
                            self.ws.delete_cols(13, self.ws.max_column - 13)
            else:
                if int(head_split[1][:1]) == 0:
                    self.quarter = datetime.datetime(year=int(head_split[2]), month=int(head_split[1][1:]), day=int(head_split[0])) + relativedelta(months=+3, days=+1)
                    #self.quarter = datetime.datetime.now()
                    if self.quarter.strftime("%d/%m/%Y") <= self.date:
                        flag = mail(self.toaddr)
                        if flag == 1:
                            self.ws.delete_cols(13, self.ws.max_column - 13)
                        else:
                            self.backup()
                            self.ws.delete_cols(13, self.ws.max_column - 13)
                else:
                    self.quarter = datetime.datetime(year=int(head_split[2]), month=int(head_split[1]), day=int(head_split[0])) + relativedelta(months=+3, days=+1)
                    if self.quarter.strftime("%d/%m/%Y") <= self.date:
                        flag = mail(self.toaddr)
                        if flag == 1:
                            self.ws.delete_cols(13, self.ws.max_column - 13)
                        else:
                            self.backup()
                            self.ws.delete_cols(13, self.ws.max_column - 13)
           
        '''
        self.toaddr = "ajay66728@gmail.com"
        #should be -2 because we are calling quarterly at first
        if (self.ws.max_column >= 102):
            flag = mail(self.toaddr)
            if flag == 1:
                self.ws.delete_cols(13, self.ws.max_column - 12)
            else:
                self.backup()
                self.ws.delete_cols(13, self.ws.max_column - 13)
        self.wb.save(self.filename)
        '''

        except:
            pass
        '''

    def client_report(self, policy, name):
        flag = 0
        index = 1
        policy_info = policy.lower()
        name_info = name.lower()
        report = dict()
        client = {}

        for row in range(2, self.ws.max_row+1):
            index += 1
            if name_info == "" and policy_info == "":
                flag = 0

            elif name_info != "" and policy_info == "":
                cell_name = self.ws.cell(row=row, column=2)
                if name_info == str(cell_name.value).lower():
                    flag = 1
                    break

            elif name_info == "" and policy_info != "":
                cell_policy = self.ws.cell(row=row, column=5)
                if policy_info == str(cell_policy.value).lower():
                    flag = 1
                    break

            else:
                cell_name = self.ws.cell(row=row, column=2)
                cell_policy = self.ws.cell(row=row, column=5)
                if name_info == str(cell_name.value).lower():
                    if policy_info == str(cell_policy.value).lower():
                        flag = 1
                        break
        name = self.ws.cell(row=index, column=2)
        client["name"] = name.value
        total = 0
        if flag == 1:
            for i in range(13, self.ws.max_column+1):
                head_obj = self.ws.cell(row=1, column=i)
                cell_obj = self.ws.cell(row=index, column=i)
                try:
                    total = total + int(self.ws.cell(row=index, column=i).value)
                except TypeError:
                    total = total + 0
                report.update({head_obj.value: cell_obj.value})
            self.report_filename = "C:/Users/AJAY LIC/Documents/LIC/Report.xlsx"
            book = Workbook()
            sheet = book.create_sheet(index=0, title=name.value)
            head = ['Date', 'Collection']
            sheet.append(head)
            row = 2
            for dict_keys, dict_values in report.items():
                sheet.cell(row=row, column=1).value = dict_keys
                sheet.cell(row=row, column=2).value = dict_values
                row += 1
            book.save(filename=self.report_filename)

            return report, client, total
        else:
            return "No Customer report available with the given data"

    def open_report(self):
        #os.system("start C:/Users/AJAY LIC/Documents/LIC/Report.xlsx")
        os.startfile('C:/Users/AJAY LIC/Documents/LIC/Report.xlsx')

    def read_entries(self, row):
        row = row + 1
        cell_name = self.ws.cell(row=row, column=2).value
        cell_address = self.ws.cell(row=row, column=3).value
        cell_daily = self.ws.cell(row=row, column=10).value
        cell_balance = self.ws.cell(row=row, column=12).value
        return cell_name, cell_address, cell_daily, cell_balance

    def datetime(self, date):
        if int(date[0][:1]) == 0:
            if int(date[1][:1]) == 0:
                date_time = datetime.datetime(year=int(date[2]), month=int(date[1][1:]), day=int(date[0][1:]))
            else:
                date_time = datetime.datetime(year=int(date[2]), month=int(date[1]), day=int(date[0][1:]))
        else:
            if int(date[1][:1]) == 0:
                date_time = datetime.datetime(year=int(date[2]), month=int(date[1][1:]), day=int(date[0]))
            else:
                date_time = datetime.datetime(year=int(date[2]), month=int(date[1]), day=int(date[0]))
        return date_time

    def view_previous(self, pre_date):
        date_split = pre_date.split("-")
        date_compare = self.datetime(date_split)
        row = 1
        pre_col = 0
        flag = 0
        total = 0
        details = list()
        entry = list()
        date_now = datetime.datetime.now().strftime("%d/%m/%Y")
        now_split = date_now.split("/")
        now = self.datetime(now_split)
        if date_compare > now:
            return "Cannot enter future date"
        else:
            for col in range(self.ws.max_column, 12, -1):
                cell_object = self.ws.cell(row=row, column=col).value
                cell_split = cell_object.split("/")
                cell_compare = self.datetime(cell_split)
                if cell_compare == date_compare:
                    pre_col = col
                    flag = 1
                    break

        if flag:
            for i in range(2, self.ws.max_row + 1):
                cell_s_no = self.ws.cell(row=i, column=1).value
                entry.append(cell_s_no)
                cell_name = self.ws.cell(row=i, column=2).value
                entry.append(cell_name)
                cell_address = self.ws.cell(row=i, column=3).value
                entry.append(cell_address)
                cell_policy = self.ws.cell(row=i, column=5).value
                entry.append(cell_policy)
                cell_collection = self.ws.cell(row=i, column=pre_col).value
                entry.append(cell_collection)
                try:
                    total = total + int(self.ws.cell(row=i, column=pre_col).value)
                except TypeError:
                    total = total + 0
                details.append(copy.deepcopy(entry))
                del entry[:]

            return details, total
        else:
            return "No collection available on the date entered. Choose another date"

    def previous_date(self, pre_date):
        date_split = pre_date.split("-")
        date_compare = self.datetime(date_split)
        row = 1
        flag = 0
        self.column = 0
        if self.ws.max_column == 12:
            self.column = 13
            flag = 1
        else:
            last_date = date.today().strftime("%d/%m/%Y")
            last_split = last_date.split("/")
            last_compare = self.datetime(last_split)
            if date_compare > last_compare:
                return flag
            for col in range(self.ws.max_column, 12, -1):
                cell_object = self.ws.cell(row=row, column=col).value
                cell_split = cell_object.split("/")
                cell_compare = self.datetime(cell_split)
                if cell_compare == date_compare:
                    self.column = col
                    flag = 0
                    break
                elif date_compare < cell_compare:
                    self.column = col
                    flag = 1
                else:
                    self.column = col+1
                    flag = 1
                    break
        if flag:
            self.ws.insert_cols(self.column)
            self.ws.cell(row=1, column=self.column).value = date_compare.strftime("%d/%m/%Y")
            self.wb.save(self.filename)
        return 1

    def delete_column(self):
        try:
            self.column
            #self.column.destroy()
            del self.column
        except Exception as e:
            pass

    def add_daily_collection(self, row_no, daily_collection):
        # calling update_daily_client function to create a new column with today's date as heading
        self.quarterly()
        try:
            try:
                self.column
                column = self.column
            except:
                self.update_daily_client()
                column = self.ws.max_column
            #daily_collection_before = int(self.ws.cell(row=row_no, column=self.ws.max_column).value)
            daily_collection_before = int(self.ws.cell(row=row_no, column=column).value)
        except TypeError:
            daily_collection_before = 0
        updated_daily_collection = daily_collection_before + int(daily_collection)
        self.ws.cell(row=int(row_no), column=10).value = daily_collection
        balance = self.ws.cell(row=int(row_no), column=12).value
        balance = int(balance) - int(daily_collection)
        flag = 0
        if balance == 0:
            self.ws.cell(row=row_no, column=12).value = int(self.ws.cell(row=int(row_no), column=9).value)
            flag = 1
        else:
            self.ws.cell(row=row_no, column=12).value = balance
        # updating the new column and corresponding entry row with today's daily collection
        self.ws.cell(row=row_no, column=column).value = updated_daily_collection
        
        # calling quarterly to check for 90 days and do a clean up of daily collection columns
        #self.quarterly()
        self.wb.save(self.filename)

        return flag

    def subtract_daily_collection(self, row_no, daily_collection):
        # calling update_daily_client function to create a new column with today's date as heading
        try:
            try:
                self.column
                column = self.column
            except:
                self.update_daily_client()
                column = self.ws.max_column
            daily_collection_before = int(self.ws.cell(row=row_no, column=column).value)
        except TypeError:
            daily_collection_before = 0
        daily_collection_update = daily_collection_before - int(daily_collection)
        self.ws.cell(row=int(row_no), column=10).value = daily_collection
        balance = self.ws.cell(row=int(row_no), column=12).value
        balance = int(balance) + int(daily_collection)
        self.ws.cell(row=row_no, column=12).value = balance
        # updating the new column and corresponding entry row with today's daily collection
        self.ws.cell(row=row_no, column=column).value = daily_collection_update
        self.wb.save(self.filename)

